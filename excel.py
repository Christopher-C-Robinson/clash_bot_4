import openpyxl as xl
from utilities import *
from admin import *

def set_last_row():
    file = ROOT_DIR + "/excel/log.xlsx"
    try:
        log_wb = xl.load_workbook(file)
        log_ws = log_wb["Sheet1"]
        admin.excel_last_row_log = len(list(log_ws.iter_rows()))
        log_ws.cell(1, 1).value = "Log"
        log_wb.save(file)
        log_wb.close()
        print("Last row:", admin.excel_last_row_log)
    except:
        print()
        print("You've got Excel open.")
        # time.sleep(60)

def log(admin_mode, account_mode, job, account, duration):
    file = ROOT_DIR + "/excel/log.xlsx"
    time = datetime.now()
    print("Log:", time, admin_mode, account_mode, job, account, duration)
    log_wb = xl.load_workbook(file)
    log_ws = log_wb["Sheet1"]
    row = admin.excel_last_row_log + 1

    try:
        values = [time, admin_mode, account_mode, job, account, duration]
        for count, value in enumerate(values, 1):
            log_ws.cell(row, count).value = value
        log_wb.save(file)
        log_wb.close()
        admin.excel_last_row_log += 1
    except:
        print("Couldn't update log")

def excel_write_rows(file, sheet, start_row, values):
    try:
        file = ROOT_DIR + "/excel/" + file + ".xlsx"
        wb = xl.load_workbook(file)
        ws = wb[str(sheet)]
        row = start_row
    except:
        print("Couldn't update log")
        return

    for row_x in ws['A4:E100']:
        for cell in row_x:
            cell.value = None

    for single_row in values:
        try:
            for count, value in enumerate(single_row, 1):
                ws.cell(row, count).value = value
            row += 1
        except:
            print("Couldn't update log")

    try:
        wb.save(file)
        wb.close()
    except:
        pass




# time = timedelta(hours=1)

# excel_write_rows(file="remaining_time", sheet=1, row=4, values=["archer", 5, 6, time, 3, 3 * time])

