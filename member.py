from nav import *
import openpyxl as xl

war_directory = "images/members/names/"
member_files = os.listdir(war_directory)
members = []

CHAT_NAME = (158, 132, 160, 770)

class Member():
    def __init__(self, name):
        self.name = name
        file = war_directory + name + ".png"
        # print(file)
        self.image = Image(file, name=self.name, threshold=0.95, no_of_regions=1, region_limit=[400, 240, 300, 700])
        # print("Image name:", self.image.name)
        # self.image.show()
        self.stars = None
        members.append(self)

    def __str__(self):
        return self.name

    def get_stars(self):
        if self.stars: return
        result, loc, rect = self.image.find_detail(show_image=False)
        # print(self, result)
        if result > self.image.threshold:
            y = rect[1] + 10
            star_rect = [1540, y, 140, 40]
            screen = get_screenshot(region=star_rect, colour=0)
            # show(screen)
            self.stars = war_stars.read_one_screen(screen)
        # print(self, self.stars)


for file in member_files:
    name = file[0:-4]
    new = Member(name)
    # print(new.name)

ma = next((x for x in members if x.name == 'max'), None)
ka = next((x for x in members if x.name == 'ka'), None)
abood = next((x for x in members if x.name == 'abood'), None)
# print(abood)

def move_war_screen(dir, distance):
    time = 0.1
    if distance > 200:
        time = distance * 2 / 450
    time = distance * 2 / 450

    if dir == "up":
        pag.moveTo(1726, 900, 0.3)
        pag.dragTo(1726, 900 - distance, time, button="left")

def war_save_name_image(rect):
    y = rect[1] - 10
    image_rect = [447, y, 140, 40]
    image_rect_large = [437, y-10, 160, 60]
    file_number = get_next_member_number()
    filename = f"images/members/names/{file_number}.png"
    screen = get_screenshot(region=image_rect_large, colour=0)
    already_exists = False
    for member in members:
        if not already_exists:
            bool, val = member.image.find_screen(screen, show_image=False, return_result=True)
            if bool:
                print("Member already exists:", member, val)
                already_exists = True
    if not already_exists:
        result = get_screenshot(region=image_rect, colour=1)
        cv2.imwrite(filename, result)
    # show(result)

def war_team_find_stars():
    rects = i_war_star.find_many()
    return rects

def war_team_scroll_up():
    star_x, star_y = 1600, 900
    pag.moveTo(star_x, star_y, 0.3)
    pag.dragTo(star_x, star_y - 550, 1.5, button="left")

def get_next_member_number():
    dir = "images/members/names"
    existing_names = os.listdir(dir)
    max_no = 1
    for name in existing_names:
        name_ex = name[:-4]
        if name_ex.isnumeric():
            max_no = max(int(name_ex), max_no)
    return max_no + 1

def get_stars():
    goto(l_war_team)
    for x in range(10):
        war_team_scroll_up()
        for member in members:
            member.get_stars()

def save_member_images():
    goto(l_war_team)
    for x in range(10):
        war_team_scroll_up()
        time.sleep(2)
        rects = war_team_find_stars()
        for rect in rects:
            war_save_name_image(rect)

def save_stars():
    file = "excel/war_stars.xlsx"
    workbook = xl.load_workbook(file)
    today_date = datetime.now().strftime("%Y-%m-%d")

    # Check if a sheet with today's date exists
    if today_date in workbook.sheetnames:
        sheet_to_delete = workbook[today_date]
        workbook.remove(sheet_to_delete)

    sheet = workbook.create_sheet("MyNewSheet")
    sheet.title = datetime.now().strftime("%Y-%m-%d")

    count = 1
    for member in members:
        if member.stars is not None:
            sheet.cell(count, 1).value = member.name
            sheet.cell(count, 2).value = member.stars
            count += 1
    workbook.save(file)

if __name__ == "__main__":
    # Add new members
    # save_member_images()

    # Save stars
    get_stars()
    save_stars()
    goto(pycharm)


