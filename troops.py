import openpyxl as xl
from member import *
from nav import *

troops = []
just_troops = []
spells = []
siege_troops = []
troop_files = os.listdir("images/troops")
troop_directory = 'images/troops/'
slide_position = 1

def get_image_from_file(troop, style):
    file = f'{troop}_{style}.png'
    if style == "": file = f'{troop}.png'
    image = None
    if file in troop_files:
        image = cv2.imread(troop_directory + file, 0)
    else:
        print(f"Adding troop image: could not find '{file}'")
    return image

def get_image(troop, style):
    object = next((x for x in troops if x.name == troop), None)
    if style == "army": return object.army
    if style == "train": return object.train
    if style == "donate1": return object.donate1
    if style == "donate2": return object.donate2
    if style == "attack": return object.attack
    if style == "research": return object.research
    return

def get_super_troop(troop):
    goto(main)
    hold_key("d", 0.5)
    hold_key("s", 0.5)
    hold_key("d", 0.5)
    troop_image = i_super_barb
    if troop == super_minion: troop_image = i_super_minion
    images = [i_boost, i_boost_on, troop_image, i_activate, i_000, i_000v2, i_red_cross_super_troops]
    for image in images:
        time.sleep(0.5)
        if image == i_super_minion:
            pag.moveTo(1070, 928)
            pag.dragTo(1070, 287, .6)
        print(image, image.find_detail())
        result = image.click()
        if not result:
            if image == i_000: i_00.click()
            if image == i_000v2: i_00v2.click()




class Troop():
    def __init__(self, name, type, slide, castle_slide, training_time, donate_bool, donate_preference, donations, donation_count):
        self.name = name
        self.type = type
        self.slide = slide
        self.castle_slide = castle_slide
        self.training_time = training_time
        self.donate_bool = donate_bool
        self.donate_preference = donate_preference
        self.currently_training = False
        if type not in ["hero", "clan"]:
            train_directory = "images/troops/train/"
            train_files = os.listdir(train_directory)
            if f"{self.name}_train.png" in train_files:
                self.i_train = Image(name=f"i_{name}_train", file=f'{train_directory}{name}_train.png', no_of_regions=1, type="train")
            else:
                self.i_train = Image(name=f"i_{name}_train", file=f'{troop_directory}{name}_train.png', no_of_regions=1, type="train")

            self.i_army = Image(name=f"i_{name}_army", file=f'{troop_directory}{name}_army.png', no_of_regions=1, type="army", threshold=0.72)
            self.i_training = Image(name=f"i_{name}_training", file=f'{troop_directory}{name}_training.png', no_of_regions=1, type="training")
            self.i_donate1 = Image(name=f"i_{name}_donate1", file=f'{troop_directory}{name}_donate1.png', no_of_regions=1, type="donate1")
            self.i_donate2 = Image(name=f"i_{name}_donate2", file=f'{troop_directory}{name}_donate2.png', no_of_regions=2, type="donate2")
            self.i_castle = Image(name=f"i_{name}_castle", file=f'{troop_directory}{name}_castle.png', no_of_regions=1, type="castle")
            if "super" not in name:
                self.i_research = Image(name=f"i_{name}_research", file=f'{troop_directory}{name}_research.png', no_of_regions=1, type="research")
        else :
            self.i_army = None
            self.i_train = None
            self.i_training = None
            self.i_donate1 = None
            self.i_donate2 = None
            self.i_castle = None
            self.research = None

        self.i_attack = Image(name=f"i_{name}_attack", file=f'{troop_directory}{name}_attack.png')

        self.donations = donations
        self.donation_count = donation_count
        self.super_troop = False
        if "super" in self.name:
            self.super_troop = True

        troops.append(self)
        if type == "troop": just_troops.append(self)
        if type == "spell": spells.append(self)
        if type == "siege": siege_troops.append(self)

    def __str__(self):
        return self.name

    def in_castle(self):
        val, loc, rect = find(self.i_army.image, get_screenshot(CASTLE_TROOPS), show_image=False)
        return val > 0.8

    def start_train(self, count, account, move_to_start=False):
        if count == 0: return
        print("Start train:", self, count)
        global slide_position
        if self.type == "troop":
            goto(troops_tab)
            slide(slide_position, self.slide)
        if self.type == "spell":
            goto(spells_tab)
        if self.type == "siege":
            # print("Account has siege", account.has_siege, self.type)
            if not account.has_siege: return
            goto(siege_tab)
            slide(slide_position, self.slide)
        if self.i_train is None:
            print("Can't train - no image")
            return

        val, loc, rect = find(self.i_train.image, get_screenshot(TRAIN_RANGE))
        # screen = get_screenshot(TRAIN_RANGE)
        # show(screen)
        # show(self.i_train.image)
        # print("Troop train", val)
        if val < 0.8 and self.super_troop and account.th >= 11:
            get_super_troop(self)
            goto(troops_tab)

        if val > 0.8:
            for x in range(count):
                click_rect(rect, region=TRAIN_RANGE)
        # val, outcome = click(self.train, region=TRAIN_RANGE, show_image=False)
        if move_to_start:
            move_to_queue_start(self)

        return self.training_time * count

    def delete(self, count):
        # print("Troop delete")
        if self.i_army.image is None: return
        goto(army_tab)
        i_army_edit.click()
        val, loc, rect = find(self.i_army.image, get_screenshot(ARMY_SPELLS_EXISTING))
        rect_adj = [rect[0] + ARMY_EXISTING[0], rect[1] + ARMY_EXISTING[1], rect[2], rect[3]]
        spot = pag.center(rect_adj)
        for x in range(count):
            pag.click(spot)
        time.sleep(0.1)
        i_army_okay.click()
        time.sleep(0.1)
        i_army_okay2.click()

levels_filename = 'C:/Users/darre/OneDrive/Darren/clash_bot/levels.xlsx'

def load_troops():
    wb = xl.load_workbook(levels_filename)
    sheet = wb['Troops']
    for row in range(2, sheet.max_row + 1):
        name = sheet.cell(row, 1).value
        type = sheet.cell(row, 2).value
        slide = sheet.cell(row, 3).value
        castle_slide = sheet.cell(row, 4).value
        training_time = sheet.cell(row, 5).value
        donate_bool = sheet.cell(row, 6).value
        donate_preference = sheet.cell(row, 7).value
        donations = sheet.cell(row, 8).value
        donation_count = sheet.cell(row, 9).value
        if name == "goblin":
            donate_bool = True
            donate_preference = 10
            donation_count = 10
            donations = 1

        # print("Creating troop:", name)
        Troop(name=name, type=type, slide=slide, castle_slide=castle_slide, training_time=training_time, donate_bool=donate_bool,
              donate_preference=donate_preference, donations=donations, donation_count=donation_count)

# clan = Troop(name="clan", type="clan", slide=1, castle_slide=1, training_time=0, donate_bool=False,
               # donate_preference=100, donations=0, donation_count=0)

def get_troop(string):
    return next((x for x in troops if x.name == string), None)

load_troops()
clan = next((x for x in troops if x.name == 'clan'), None)

barb = next((x for x in troops if x.name == 'barb'), None)
archer = next((x for x in troops if x.name == 'archer'), None)
goblin = next((x for x in troops if x.name == 'goblin'), None)
giant = next((x for x in troops if x.name == 'giant'), None)
wizard = next((x for x in troops if x.name == 'wizard'), None)
bomber = next((x for x in troops if x.name == 'bomber'), None)
bloon = next((x for x in troops if x.name == 'bloon'), None)
dragon = next((x for x in troops if x.name == 'dragon'), None)
baby_drag = next((x for x in troops if x.name == 'baby_dragon'), None)
edrag = next((x for x in troops if x.name == 'edrag'), None)
minion = next((x for x in troops if x.name == 'minion'), None)
hog = next((x for x in troops if x.name == 'hog'), None)
golem = next((x for x in troops if x.name == 'golem'), None)
witch = next((x for x in troops if x.name == 'witch'), None)
lava_hound = next((x for x in troops if x.name == 'lava_hound'), None)
ice_golem = next((x for x in troops if x.name == 'ice_golem'), None)
headhunter = next((x for x in troops if x.name == 'headhunter'), None)
titan = next((x for x in troops if x.name == 'titan'), None)

super_barb = next((x for x in troops if x.name == 'super_barb'), None)
super_minion = next((x for x in troops if x.name == 'super_minion'), None)
# super_minion.slide = 2

lightening = next((x for x in troops if x.name == 'lightening'), None)
heal = next((x for x in troops if x.name == 'heal'), None)
rage = next((x for x in troops if x.name == 'rage'), None)
freeze = next((x for x in troops if x.name == 'freeze'), None)
poison = next((x for x in troops if x.name == 'poison'), None)
skeleton = next((x for x in troops if x.name == 'skeleton'), None)
clone = next((x for x in troops if x.name == 'clone'), None)
quake = next((x for x in troops if x.name == 'quake'), None)

queen = next((x for x in troops if x.name == 'queen'), None)
king = next((x for x in troops if x.name == 'king'), None)
warden = next((x for x in troops if x.name == 'warden'), None)
champ = next((x for x in troops if x.name == 'champ'), None)

ram = next((x for x in troops if x.name == 'ram'), None)
blimp = next((x for x in troops if x.name == 'blimp'), None)
slammer = next((x for x in troops if x.name == 'slammer'), None)
log_thrower = next((x for x in troops if x.name == 'log_thrower'), None)
flinger = next((x for x in troops if x.name == 'flinger'), None)

troops.sort(key=lambda x: x.donate_preference, reverse=False)

def troop_str(troops):
    string = ""
    for x in troops:
        try:
            string += x.name + ", "
        except:
            pass
    return string[0:-1]


def delete_a_troop():
    val, loc, rect = find_cv2("remove_troops", DELETE_2_REGION)
    center = pag.center(rect)
    pag.click(center)

def make_room(troop):
    count = 0
    colour = check_troop_colour_train(troop)
    print("Make room:", colour)
    while not colour and count < 6:
        delete_a_troop()
        colour = check_troop_colour_train(troop)
        count += 1

def check_troop_colour_train(troop):
    if troop.train is None:
        print("No training image")
        return False
    val, loc, rect = find(troop.train, get_screenshot(TRAIN_RANGE), troop.name)
    rect_adj = [rect[0] + TRAIN_RANGE[0], rect[1] + TRAIN_RANGE[1], rect[2], rect[3], ]
    colour = check_colour_rect(rect_adj, show_image=False)
    return colour

def move_to_queue_start(troop):
    print("Moving to queue start")
    for _ in range(2):
        val, loc, rect = find(troop.training, get_screenshot(BACKLOG))
        if val > 0.70:
            a = BACKLOG[0] + pag.center(rect)[0], BACKLOG[1] + pag.center(rect)[1] + 10
            b = BACKLOG[0] + BACKLOG[2] + 10, a[1] + 10
            if a[0] < 1600: drag(a, b)

def drag(a, b):
    x0, y0 = a
    x1, y1 = b
    pag.moveTo(x0, y0, 0.3)
    pag.dragTo(x1, y1, 1, button="left")
    time.sleep(1)

def slide(slide_pos, slide_pos_target):
    # print("Sliding:", slide_pos, slide_pos_target)
    time.sleep(0.2)
    if slide_pos == slide_pos_target:
        return slide_pos

    if slide_pos < slide_pos_target:
        start_x = 1500
        end_x = 500
        # slide_pos = min(slide_pos + 1, 2)
    else:
        start_x = 500
        end_x = 1500
        # slide_pos = min(slide_pos - 1, 1)

    dur = 0.3
    # print(slide_pos, slide_pos_target)
    # print((slide_pos - slide_pos_target))
    # print(abs(slide_pos - slide_pos_target))
    for x in range(abs(slide_pos - slide_pos_target)):
        print("Slide loop:", x)
        pag.moveTo(start_x, 660, dur)
        pag.dragTo(end_x, 660, dur)
        time.sleep(1)
    slide_pos = slide_pos_target
    return slide_pos

def merge_troop_regions():
    for image_type in ["donate1", "donate2"]:
        all_images = [x for x in images if image_type in x.name]
        for image in all_images: print(image.name)

        for image1 in all_images:
            for image2 in all_images:
                if image1 == image2: continue
                for region in image1.regions:
                    if region not in image2.regions:
                        image2.save_region(region)


# log_thrower.i_donate2.show_regions()

# for x in troops:
#     print(x)
#     print(x.i_donate2.image)
#     if x.i_donate2.image:
#         show(x.i_donate2)

